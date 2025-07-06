import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def col(letter):
    return column_index_from_string(letter) - 1

def classify_calculation_method(file_path):
    sheet_name = '表3.活動數據'

    # 1. Read from row 4 (skip first 3 rows)
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=3, engine='openpyxl')

    # 2. Column mappings using Excel letters
    col_source = col('C')     # 排放源
    col_category = col('F')   # 排放類別
    col_unit = col('O')       # 單位
    col_type = col('A')       # 類別 or type1 in SQLite db
    col_activity = col('B')  # 活動/設備/type2 in SQLite db

    # 3. Define flexible rule set
    rules = [
        #SCOPE 1 - DIRECT EMISSIONS - 範疇1
        #STATIONARY COMBUSTION - 固定源
        {
            'match': lambda row: row[col_unit] in ['公秉', '千公秉','千立方公尺', '立方公尺', '公斤','公升', '公噸'] 
                        and row[col_category] in ['固定源'],
            'formula': lambda row: f"{row[col_source]}使用量×排放係數×GWP值"
        },
        #MOBILE COMBUSTION - 移動源
        {
            'match': lambda row: row[col_unit] in ['公秉', '千公秉','千立方公尺', '立方公尺', '公斤','公升'，'公噸'] 
                        and row[col_category] in ['移動源'],
            'formula': lambda row: f"{row[col_source]}使用量×排放係數×GWP值"
        },
        #FUGITIVE EMISSIONS - 逸散排放
        #Refrigerants 冷媒
        {
            'match': lambda row: 'R' in str(row[col_source])
                    and row[col_category] in ['逸散排放'],
            'formula': lambda row: f"冷媒設備填充量/規格量×設備逸散因子×排放係數×GWP值\n設備逸散因子來源參考2006 IPCC Guidelines for National Greenhouse Gas Inventories, volume 3, chapter7, table 7.9，採用中間值作為逸散因子"
        },
        #Septic Tank 化糞池
        {
            'match': lambda row: row[col_type] in ['化糞池']
                    and row[col_category] in ['逸散排放'],
            'formula': lambda row: "員工數×對應工作天數與時數之排放係數×GWP值"
        },
        #Fire Extinguisher 滅火器
        {
            'match': lambda row: row[col_category] in ['逸散排放']
                        and row[col_type] in ['消防設施']
                        and '滅火' in str(row[col_activity]),
            'formula': lambda row: "填充使用重量×排放係數×GWP值"
        },
        #Other CO2
        {
            'match': lambda row: 'CO2' in str(row[col_source]) 
                        and row[col_category] in ['逸散排放']
                        and row[col_type] in ['其他設施']
                        and row[col_activity] in ['其他設施'],
            'formula': lambda row: "質量平衡法"
        },

        #SCOPE 2 - INDIRECT EMISSIONS - 範疇2
        #Purchased Electricity 外購電力
        {
            'match': lambda row: row[col_category] in ['外購電力'],
            'formula': lambda row: f"電力使用度數×排放係數×GWP值\n(外購電力排放係數採用能源局公告之112年電力排碳係數0.494公斤CO₂e/度計算)"
        },
        #Purchased Steam 外購蒸汽 ASKAMY about the formula or see SDC
        {
            'match': lambda row: '外購蒸汽' in str(row[col_source])
                        or row[col_activity] in ['外購蒸汽'],
            'formula': lambda row: "使用量×排放係數×GWP值"
        },
        #Self-Sustained Electricity 自發自用 ASKAMY about the formula or see SDC
        {
            'match': lambda row: '自發自用' in str(row[col_source])
                        or row[col_activity] in ['自發自用']
                        or row[col_category] in ['自發自用'],
            'formula': lambda row: "使用量×排放係數×GWP值"
        },

        #CATEGORY 1 - PURCHASED GOODS & SERVICES - 類別1 - 採購商品與服務
        #Weight-Dependent
        {
            'match': lambda row: ('公噸') in str(row[col_unit])
                        and row[col_category] in ['採購商品與服務']
                        and '自來水' not in str(row[col_source]),
            'formula': lambda row: "採購重量×排放係數×GWP值"
        },
        #Volume-Dependent
        {
            'match': lambda row: ('公秉' or '立方公尺') in str(row[col_unit])
                        and row[col_category] in ['採購商品與服務']
                        and '自來水' not in str(row[col_source]),
            'formula': lambda row: "採購體積×排放係數×GWP值"
        },
        #Price-Dependent (EEIO) ASKAMY about the formula
        {
            'match': lambda row: ('元' or '金') in str(row[col_unit])
                        and row[col_category] in ['採購商品與服務'],
            'formula': lambda row: "採購金額×別幣到2022美金轉換率×排放係數×GWP值"
        },
        #Water Supply - 自來水 ASKAMY about the criteria
        {
            'match': lambda row: ('自來水') in str(row[col_source])
                        and row[col_category] in ['採購商品與服務'],
            'formula': lambda row: "用水量x排放係數xGWP值"
        },
        
        #CATEGORY 2 - CAPITAL GOODS - 類別2 - 資本財
        {
            'match': lambda row: ('元' or '金') in str(row[col_unit])
                        and row[col_category] in ['資本財'],
            'formula': lambda row: "採購金額×別幣到2022美金轉換率×排放係數×GWP值"
        },

        #CATEGORY 3 - FUEL- & ENERGY- RELATED EMISSIONS - 類別3 - 與燃料和能源相關的活動
        #Purchased Electricity 電力
        {
            'match': lambda row: row[col_unit] in ['公秉', '千公秉','千立方公尺', '立方公尺', '公斤','公升','公噸','度', '千度'] 
                        and row[col_category] in ['與燃料和能源相關的活動']
                        and '電力' in str(row[col_source]),
            'formula': lambda row: "電力使用量×排放係數×GWP值"
        },
        #Natural Gas 天然氣
        {
            'match': lambda row: row[col_unit] in ['公秉', '千公秉','千立方公尺', '立方公尺', '公斤','公升','公噸','度', '千度'] 
                        and row[col_category] in ['與燃料和能源相關的活動']
                        and '天然氣' in str(row[col_source]),
            'formula': lambda row: "天然氣使用量×排放係數×GWP值"
        },
        #Petroleum Gas 汽油
        {
            'match': lambda row: row[col_unit] in ['公秉', '千公秉','千立方公尺', '立方公尺', '公斤','公升','公噸','度', '千度'] 
                        and row[col_category] in ['與燃料和能源相關的活動']
                        and '汽油' in str(row[col_source]),
            'formula': lambda row: "汽油使用量×排放係數×GWP值"
        },
        #Diesel 柴油
        {
            'match': lambda row: row[col_unit] in ['公秉', '千公秉','千立方公尺', '立方公尺', '公斤','公升','公噸','度', '千度'] 
                        and row[col_category] in ['與燃料和能源相關的活動']
                        and '柴油' in str(row[col_source]),
            'formula': lambda row: "柴油使用量×排放係數×GWP值"
        },
        # Add more rules here as needed
    ]

    # 4. Apply rules
    def get_calculation_method(row):
        for rule in rules:
            try:
                if rule['match'](row):
                    return rule['formula'](row)
            except Exception as e:
                return f"錯誤: {e}"
        return ""

    df['計算方式'] = df.apply(get_calculation_method, axis=1)

    # 5. Write back to Excel
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    start_row = 4
    new_col_index = ws.max_column + 1
    ws.cell(row=3, column=new_col_index, value='計算方式')  # Header

    for i, value in enumerate(df['計算方式'], start=start_row):
        ws.cell(row=i, column=new_col_index, value=value)

    wb.save(file_path)
    print("✅ 計算方式 column added with Excel-style column mapping.")
